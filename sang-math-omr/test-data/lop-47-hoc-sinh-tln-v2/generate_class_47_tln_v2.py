#!/usr/bin/env python3
"""Generate a deterministic 47-student TLN-v2 OMR regression bundle.

The bundle exercises Vietnamese left-to-right short answers on the
12-4-6 A5-landscape sheet.  It deliberately covers:

* ``0`` in the first TLN column;
* ``-2`` in the first two columns;
* ``4,5`` in the first three columns;
* ``-2,3`` in all four columns;
* positive integers and decimal answers from one to four characters.

The checked-in coordinate JSON predates the extra zero bubble in TLN column
one.  ``upgrade_template`` mirrors ``js/tln_codec.js`` by appending that point
at legacy index 10 without moving any existing coordinate.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import random
import re
import subprocess
import tempfile
from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw


SEED = 20260716
STUDENT_COUNT = 47
EXAM_CODE = "0303"
TEMPLATE_ID = "12-4-6ngang"
PPI = 144
COORDINATE_PPI = 72
SCALE = PPI / COORDINATE_PPI
FILL_RADIUS = int(3.8 * SCALE)
MAX_PDF_BYTES = 25 * 1024 * 1024

HERE = Path(__file__).resolve().parent
OMR_ROOT = HERE.parents[1]
REPO_ROOT = OMR_ROOT.parent
TEMPLATES_JSON = OMR_ROOT / "templates.json"
TYPST_TEMPLATE = OMR_ROOT / "templates" / "12-4-6ngang.typ"

PDF_PATH = HERE / "lop-47-hoc-sinh-tln-v2.pdf"
ANSWER_JSON_PATH = HERE / "dap-an-ma-de-0303.json"
ANSWER_CSV_PATH = HERE / "dap-an-ma-de-0303.csv"
ANSWER_XLSX_PATH = HERE / "dap-an-ma-de-0303.xlsx"
STUDENT_DATA_PATH = HERE / "du-lieu-47-hoc-sinh.json"
EXPECTED_PATH = HERE / "ket-qua-ky-vong.csv"
HASH_PATH = HERE / "SHA256SUMS.txt"

MCQ_OPTIONS = ("A", "B", "C", "D")
TF_OPTIONS = ("Đ", "S")
TF_LABELS = ("a", "b", "c", "d")

# Keep these indexes identical to js/tln_codec.js.
TLN_COLUMN_SYMBOLS = (
    ("-", "1", "2", "3", "4", "5", "6", "7", "8", "9", "0"),
    (",", "0", "1", "2", "3", "4", "5", "6", "7", "8", "9"),
    (",", "0", "1", "2", "3", "4", "5", "6", "7", "8", "9"),
    ("0", "1", "2", "3", "4", "5", "6", "7", "8", "9"),
)

TLN_ANSWER_KEY = {
    "17": "0",
    "18": "-2",
    "19": "4,5",
    "20": "-2,3",
    "21": "61,5",
    "22": "832",
}

TLN_VARIANTS = (
    "0",
    "1",
    "7",
    "9",
    "-2",
    "16",
    "42",
    "90",
    "4,5",
    "0,5",
    "-12",
    "832",
    "123",
    "-2,3",
    "61,5",
    "12,3",
    "1234",
    "-123",
)


def normalize_tln(value: object) -> str:
    return (
        str(value)
        .strip()
        .replace("−", "-")
        .replace("–", "-")
        .replace(".", ",")
        .replace(" ", "")
    )


def validate_tln_value(value: str) -> None:
    normalized = normalize_tln(value)
    if not 1 <= len(normalized) <= 4:
        raise ValueError(f"TLN must have 1-4 characters: {value!r}")
    for column_index, symbol in enumerate(normalized):
        if symbol not in TLN_COLUMN_SYMBOLS[column_index]:
            raise ValueError(
                f"TLN symbol {symbol!r} is invalid in column {column_index + 1}: "
                f"{value!r}"
            )


def upgrade_template(template: dict) -> dict:
    upgraded = deepcopy(template)
    for tinfo in upgraded.get("tln", {}).values():
        if (
            isinstance(tinfo, list)
            and len(tinfo) >= 2
            and len(tinfo[0]) == 10
            and len(tinfo[1]) >= 2
        ):
            # Same compatibility rule as OmrTlnCodec.upgradeTemplate().
            zero_point = [tinfo[0][0][0], tinfo[1][1][1]]
            tinfo[0].append(zero_point)
    return upgraded


def compile_blank_sheet(temp_dir: Path) -> Image.Image:
    output_pattern = temp_dir / "blank-{p}.png"
    subprocess.run(
        [
            "typst",
            "compile",
            "--pages",
            "1",
            "--ppi",
            str(PPI),
            "--root",
            str(REPO_ROOT),
            str(TYPST_TEMPLATE),
            str(output_pattern),
        ],
        check=True,
    )
    pages = sorted(temp_dir.glob("blank-*.png"))
    if len(pages) != 1:
        raise RuntimeError(f"Expected one blank A5 page, found {len(pages)}")
    with Image.open(pages[0]) as image:
        return image.convert("RGB")


def iter_coordinate_pairs(value):
    if (
        isinstance(value, list)
        and len(value) == 2
        and all(isinstance(item, (int, float)) for item in value)
    ):
        yield value
        return
    if isinstance(value, dict):
        for child in value.values():
            yield from iter_coordinate_pairs(child)
    elif isinstance(value, list):
        for child in value:
            yield from iter_coordinate_pairs(child)


def ring_angular_coverage(image: Image.Image, point: list[float]) -> int:
    """Count angular sectors containing the printed ring around a coordinate.

    A raw dark-pixel ratio is unreliable for Typst's thin anti-aliased rings.
    Angular coverage distinguishes a centered bubble (dark arcs surround the
    coordinate) from a stale profile that merely passes near one side.
    """
    gray = image.convert("L")
    cx, cy = (round(point[0] * SCALE), round(point[1] * SCALE))
    sectors = [False] * 16
    for y in range(cy - 12, cy + 13):
        for x in range(cx - 12, cx + 13):
            radius = math.hypot(x - cx, y - cy)
            if 4 <= radius <= 11 and gray.getpixel((x, y)) < 200:
                angle = math.atan2(y - cy, x - cx) + math.pi
                sector = int(angle / (2 * math.pi) * len(sectors)) % len(sectors)
                sectors[sector] = True
    return sum(sectors)


def validate_template_geometry(template: dict, blank: Image.Image) -> None:
    if blank.width <= blank.height:
        raise RuntimeError(f"Expected landscape page, got {blank.size}")

    expected_width = round(template["warp"]["width"] * SCALE)
    expected_height = round(template["warp"]["height"] * SCALE)
    if abs(blank.width - expected_width) > 1 or abs(blank.height - expected_height) > 4:
        raise RuntimeError(
            "Compiled page does not match template geometry: "
            f"page={blank.size}, warp≈({expected_width}, {expected_height})"
        )

    expected_counts = {
        "sbd": 6 * 10,
        "made": 4 * 10,
        "mcq": 12 * 4,
        "tf": 4 * 4 * 2,
        "tln": 6 * 43,
    }
    for section, expected_count in expected_counts.items():
        points = list(iter_coordinate_pairs(template[section]))
        if len(points) != expected_count:
            raise RuntimeError(
                f"Unexpected {section} coordinate count: "
                f"{len(points)} != {expected_count}"
            )
        for x, y in points:
            px, py = round(x * SCALE), round(y * SCALE)
            if not (0 <= px < blank.width and 0 <= py < blank.height):
                raise RuntimeError(f"{section} coordinate outside page: {(x, y)}")

        weakest_ring = min(ring_angular_coverage(blank, point) for point in points)
        if weakest_ring < 7:
            raise RuntimeError(
                f"{section} coordinate misses a printed bubble "
                f"(weakest angular coverage {weakest_ring}/16)"
            )

    for marker_name in ("TL", "TR", "BR", "BL"):
        x, y = template["warp"][marker_name]
        px, py = round(x * SCALE), round(y * SCALE)
        if max(blank.getpixel((px, py))) > 20:
            raise RuntimeError(f"Warp marker {marker_name} is not black at {(px, py)}")


def build_answer_key(rng: random.Random) -> dict:
    return {
        "mcq": {
            str(question): rng.choice(MCQ_OPTIONS)
            for question in range(1, 13)
        },
        "tf": {
            str(question): {
                label: rng.choice(TF_OPTIONS)
                for label in TF_LABELS
            }
            for question in range(13, 17)
        },
        "tln": dict(TLN_ANSWER_KEY),
    }


def wrong_mcq(correct: str, rng: random.Random) -> str:
    return rng.choice(tuple(option for option in MCQ_OPTIONS if option != correct))


def wrong_tln(correct: str, rng: random.Random) -> str:
    return rng.choice(tuple(value for value in TLN_VARIANTS if value != correct))


def build_student_answers(
    index: int, answer_key: dict, rng: random.Random
) -> dict:
    if index == 1:
        return deepcopy(answer_key)

    ability = 0.34 + (index - 2) * (0.61 / (STUDENT_COUNT - 2))
    ability = max(0.25, min(0.97, ability + rng.uniform(-0.08, 0.08)))

    mcq = {
        question: correct if rng.random() < ability else wrong_mcq(correct, rng)
        for question, correct in answer_key["mcq"].items()
    }
    tf = {
        question: {
            label: correct if rng.random() < ability else ("S" if correct == "Đ" else "Đ")
            for label, correct in correct_row.items()
        }
        for question, correct_row in answer_key["tf"].items()
    }
    tln = {
        question: correct if rng.random() < ability else wrong_tln(correct, rng)
        for question, correct in answer_key["tln"].items()
    }

    # A deterministic diagnostic row that exercises additional legal forms.
    if index == 2:
        tln = {
            "17": "7",
            "18": "16",
            "19": "0,5",
            "20": "61,5",
            "21": "-12",
            "22": "1234",
        }

    return {"mcq": mcq, "tf": tf, "tln": tln}


def tf_score(correct_row: dict, student_row: dict) -> float:
    correct_count = sum(
        student_row[label] == correct_row[label] for label in TF_LABELS
    )
    return (0.0, 0.1, 0.25, 0.5, 1.0)[correct_count]


def score_answers(answer_key: dict, answers: dict) -> dict:
    mcq_correct = sum(
        answers["mcq"][question] == correct
        for question, correct in answer_key["mcq"].items()
    )
    tf_points = sum(
        tf_score(correct_row, answers["tf"][question])
        for question, correct_row in answer_key["tf"].items()
    )
    tln_correct = sum(
        normalize_tln(answers["tln"][question]) == normalize_tln(correct)
        for question, correct in answer_key["tln"].items()
    )
    total = mcq_correct * 0.25 + tf_points + tln_correct * 0.5
    return {
        "mcq_correct": mcq_correct,
        "tf_points": tf_points,
        "tln_correct": tln_correct,
        "total": total,
    }


def tln_points(template: dict, local_question: str, answer: str) -> list:
    normalized = normalize_tln(answer)
    validate_tln_value(normalized)
    tinfo = template["tln"][local_question]
    points = []
    for column_index, symbol in enumerate(normalized):
        bubble_index = TLN_COLUMN_SYMBOLS[column_index].index(symbol)
        points.append(tinfo[column_index][bubble_index])
    return points


def selected_points(
    template: dict, sbd: str, exam_code: str, answers: dict
) -> list:
    points = []
    for column_index, digit in enumerate(sbd):
        points.append(template["sbd"][column_index][int(digit)])
    for column_index, digit in enumerate(exam_code):
        points.append(template["made"][column_index][int(digit)])
    for question, answer in answers["mcq"].items():
        points.append(template["mcq"][question][MCQ_OPTIONS.index(answer)])
    for global_question, row in answers["tf"].items():
        local_question = str(int(global_question) - 12)
        for label, answer in row.items():
            points.append(
                template["tf"][local_question][label][0 if answer == "Đ" else 1]
            )
    for global_question, answer in answers["tln"].items():
        local_question = str(int(global_question) - 16)
        points.extend(tln_points(template, local_question, answer))
    return points


def render_page(blank: Image.Image, points: list) -> Image.Image:
    page = blank.copy()
    draw = ImageDraw.Draw(page)
    for x, y in points:
        px, py = round(x * SCALE), round(y * SCALE)
        draw.ellipse(
            (
                px - FILL_RADIUS,
                py - FILL_RADIUS,
                px + FILL_RADIUS,
                py + FILL_RADIUS,
            ),
            fill="black",
        )
    return page


def patch_mean(
    image: Image.Image, point: list[float], radius: int = 2
) -> float:
    cx, cy = round(point[0] * SCALE), round(point[1] * SCALE)
    values = []
    for y in range(cy - radius, cy + radius + 1):
        for x in range(cx - radius, cx + radius + 1):
            values.extend(image.getpixel((x, y)))
    return sum(values) / len(values)


def validate_rendered_page(
    page: Image.Image, template: dict, points: list, answers: dict
) -> None:
    for point in points:
        if patch_mean(page, point) > 5:
            raise RuntimeError(f"Expected filled bubble is not black at {point}")

    # Verify the exact four examples that motivated TLN schema v2.
    if answers["tln"] == TLN_ANSWER_KEY:
        expected_columns = {
            "17": [0],
            "18": [0, 1],
            "19": [0, 1, 2],
            "20": [0, 1, 2, 3],
        }
        for global_question, columns in expected_columns.items():
            local_question = str(int(global_question) - 16)
            actual = tln_points(
                template, local_question, answers["tln"][global_question]
            )
            if len(actual) != len(columns):
                raise RuntimeError(
                    f"TLN {global_question} did not use {len(columns)} columns"
                )

        zero_point = template["tln"]["1"][0][10]
        if patch_mean(page, zero_point) > 5:
            raise RuntimeError("TLN answer 0 was not filled in column 1")


def answer_headers() -> list[str]:
    headers = ["made"]
    headers.extend(f"q{question}" for question in range(1, 13))
    for question in range(13, 17):
        headers.extend(f"tf{question}{label}" for label in TF_LABELS)
    headers.extend(f"tln{question}" for question in range(17, 23))
    return headers


def answer_row(answer_key: dict) -> list[str]:
    row = [EXAM_CODE]
    row.extend(answer_key["mcq"][str(question)] for question in range(1, 13))
    for question in range(13, 17):
        row.extend(
            answer_key["tf"][str(question)][label] for label in TF_LABELS
        )
    row.extend(answer_key["tln"][str(question)] for question in range(17, 23))
    return row


def write_answer_files(answer_key: dict) -> None:
    payload = {
        "v": 1,
        "meta": {
            "exam": "Bộ kiểm thử 47 học sinh - TLN trái sang phải v2",
            "subject": "Toán",
            "seed": SEED,
            "omr": {
                "id": TEMPLATE_ID,
                "name": "Toán 12-4-6 A5 ngang",
                "mcq": 12,
                "tf": 4,
                "tln": 6,
                "paper": "a5",
                "tlnSchema": 2,
            },
        },
        "keys": {EXAM_CODE: answer_key},
    }
    ANSWER_JSON_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    headers = answer_headers()
    row = answer_row(answer_key)
    with ANSWER_CSV_PATH.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerow(row)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DapAn"
    sheet.append(headers)
    sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="047857")
    for column_index, _ in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = (
            10 if column_index == 1 else 8
        )
    sheet["A2"].number_format = "@"

    guide = workbook.create_sheet("HuongDan")
    guide.append(["MỤC", "NỘI DUNG"])
    guide.append(["Mã đề", EXAM_CODE])
    guide.append(["TLN 17", "0 — tô cột 1"])
    guide.append(["TLN 18", "-2 — tô cột 1, 2"])
    guide.append(["TLN 19", "4,5 — tô cột 1, 2, 3"])
    guide.append(["TLN 20", "-2,3 — tô đủ 4 cột"])
    guide.append(["Lưu ý", "Giữ mã đề dưới dạng văn bản 0303"])
    for cell in guide[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    guide.column_dimensions["A"].width = 16
    guide.column_dimensions["B"].width = 42

    workbook.properties.creator = "SANG MATH OMR"
    workbook.properties.title = "Bộ đáp án kiểm thử TLN v2"
    workbook.save(ANSWER_XLSX_PATH)


def write_student_data(
    answer_key: dict, students: list[dict]
) -> None:
    payload = {
        "v": 1,
        "seed": SEED,
        "templateId": TEMPLATE_ID,
        "examCode": EXAM_CODE,
        "answerKey": answer_key,
        "students": students,
    }
    STUDENT_DATA_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def write_expected_results(students: list[dict]) -> None:
    fieldnames = (
        "STT",
        "SBD",
        "MaDe",
        "TLN17",
        "TLN18",
        "TLN19",
        "TLN20",
        "TLN21",
        "TLN22",
        "MCQ_Dung",
        "TF_Diem",
        "TLN_Dung",
        "Diem_Ky_Vong",
    )
    with EXPECTED_PATH.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for student in students:
            answers = student["answers"]
            score = student["expected"]
            writer.writerow(
                {
                    "STT": student["index"],
                    "SBD": student["sbd"],
                    "MaDe": EXAM_CODE,
                    **{
                        f"TLN{question}": answers["tln"][str(question)]
                        for question in range(17, 23)
                    },
                    "MCQ_Dung": score["mcqCorrect"],
                    "TF_Diem": f'{score["tfPoints"]:.2f}',
                    "TLN_Dung": score["tlnCorrect"],
                    "Diem_Ky_Vong": f'{score["score"]:.2f}',
                }
            )


def write_pdf(pages: list[Image.Image]) -> None:
    pages[0].save(
        PDF_PATH,
        "PDF",
        resolution=PPI,
        save_all=True,
        append_images=pages[1:],
        quality=88,
        optimize=True,
        title="Lop 47 hoc sinh TLN v2 - SANG MATH OMR",
        author="SANG MATH OMR test generator",
        creationDate=False,
        modDate=False,
    )


def validate_pdf() -> None:
    pdf_bytes = PDF_PATH.read_bytes()
    if len(pdf_bytes) >= MAX_PDF_BYTES:
        raise RuntimeError(
            f"Generated PDF is {len(pdf_bytes)} bytes; "
            f"limit is {MAX_PDF_BYTES}"
        )
    page_count = len(re.findall(rb"/Type /Page\b", pdf_bytes))
    if page_count != STUDENT_COUNT:
        raise RuntimeError(
            f"Expected {STUDENT_COUNT} PDF pages, found {page_count}"
        )

    info = subprocess.run(
        ["pdfinfo", str(PDF_PATH)],
        capture_output=True,
        text=True,
        check=True,
    ).stdout
    if not re.search(r"^Pages:\s+47\s*$", info, re.MULTILINE):
        raise RuntimeError("pdfinfo did not report 47 pages")
    size_match = re.search(
        r"^Page size:\s+([\d.]+) x ([\d.]+) pts",
        info,
        re.MULTILINE,
    )
    if not size_match:
        raise RuntimeError("pdfinfo did not report the page size")
    width, height = map(float, size_match.groups())
    if width <= height or abs(width - 595.5) > 1 or abs(height - 419.5) > 1:
        raise RuntimeError(f"Unexpected PDF page size: {width} x {height} pts")


def validate_coverage(students: list[dict]) -> None:
    values = {
        normalize_tln(answer)
        for student in students
        for answer in student["answers"]["tln"].values()
    }
    required = {"0", "-2", "4,5", "-2,3", "61,5", "832"}
    missing = sorted(required - values)
    if missing:
        raise RuntimeError(f"TLN regression values are missing: {missing}")
    lengths = {len(value) for value in values}
    if lengths != {1, 2, 3, 4}:
        raise RuntimeError(f"TLN lengths 1-4 not fully covered: {lengths}")
    for value in values:
        validate_tln_value(value)


def write_hash_manifest() -> None:
    paths = (
        ANSWER_JSON_PATH,
        ANSWER_CSV_PATH,
        ANSWER_XLSX_PATH,
        STUDENT_DATA_PATH,
        EXPECTED_PATH,
        PDF_PATH,
    )
    lines = [
        f"{hashlib.sha256(path.read_bytes()).hexdigest()}  {path.name}"
        for path in paths
    ]
    HASH_PATH.write_text("\n".join(lines) + "\n", encoding="ascii")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--pages-dir",
        type=Path,
        help="Optionally save the 47 rendered PNG pages for browser batch tests.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rng = random.Random(SEED)
    with TEMPLATES_JSON.open(encoding="utf-8") as source:
        template = upgrade_template(json.load(source)[TEMPLATE_ID])

    answer_key = build_answer_key(rng)
    pages = []
    students = []

    if args.pages_dir:
        args.pages_dir.mkdir(parents=True, exist_ok=True)
        for stale in args.pages_dir.glob("student-*.png"):
            stale.unlink()

    with tempfile.TemporaryDirectory(
        prefix="sang-math-omr-tln-v2-class-"
    ) as temp:
        blank = compile_blank_sheet(Path(temp))
        validate_template_geometry(template, blank)

        for index in range(1, STUDENT_COUNT + 1):
            sbd = f"{620000 + index:06d}"
            answers = build_student_answers(index, answer_key, rng)
            raw_score = score_answers(answer_key, answers)
            expected = {
                "mcqCorrect": raw_score["mcq_correct"],
                "tfPoints": raw_score["tf_points"],
                "tlnCorrect": raw_score["tln_correct"],
                "score": raw_score["total"],
            }
            points = selected_points(template, sbd, EXAM_CODE, answers)
            page = render_page(blank, points)
            validate_rendered_page(page, template, points, answers)
            pages.append(page)

            if args.pages_dir:
                page.save(
                    args.pages_dir / f"student-{index:03d}.png",
                    "PNG",
                    optimize=True,
                )

            students.append(
                {
                    "index": index,
                    "sbd": sbd,
                    "made": EXAM_CODE,
                    "answers": answers,
                    "expected": expected,
                }
            )

    validate_coverage(students)
    write_answer_files(answer_key)
    write_student_data(answer_key, students)
    write_expected_results(students)
    write_pdf(pages)
    validate_pdf()
    write_hash_manifest()

    for page in pages:
        page.close()

    print(f"Generated {STUDENT_COUNT} A5-landscape pages: {PDF_PATH}")
    print(f"PDF size: {PDF_PATH.stat().st_size} bytes")
    if args.pages_dir:
        print(f"PNG pages: {args.pages_dir}")
    print(HASH_PATH.read_text(encoding="ascii"), end="")


if __name__ == "__main__":
    main()
